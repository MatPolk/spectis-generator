"""
Generator Raportów Wojewódzkich - SPECTIS
==========================================
Aplikacja Streamlit

INSTALACJA:
pip install streamlit pandas openpyxl

URUCHOMIENIE:
streamlit run app_wojewodztwa.py

Aplikacja otworzy się automatycznie w przeglądarce.
"""

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io


def polish_sort_key(text):
    """Sortowanie z polskimi znakami"""
    replacements = {
        'ą': 'a~', 'ć': 'c~', 'ę': 'e~', 'ł': 'l~',
        'ń': 'n~', 'ó': 'o~', 'ś': 's~', 'ź': 'z~', 'ż': 'z~~',
        'Ą': 'A~', 'Ć': 'C~', 'Ę': 'E~', 'Ł': 'L~',
        'Ń': 'N~', 'Ó': 'O~', 'Ś': 'S~', 'Ź': 'Z~', 'Ż': 'Z~~'
    }
    result = text.lower()
    for pl, repl in replacements.items():
        result = result.replace(pl.lower(), repl)
    return result


# Konfiguracja strony
st.set_page_config(
    page_title="Generator Raportów Wojewódzkich",
    page_icon="📊",
    layout="wide"
)

# Tytuł
st.title("📊 Generator Raportów Wojewódzkich - SPECTIS")
st.markdown("---")

# Instrukcja
with st.expander("📖 Instrukcja użycia"):
    st.markdown("""
    ### Jak używać?
    
    1. **Wgraj plik Excel** z bazą inwestycji (przycisk poniżej)
    2. **(Opcjonalnie)** Wpisz nazwy inwestycji do wykluczenia
    3. **Kliknij "Generuj raport"**
    4. **Pobierz gotowy plik** (przycisk pojawi się automatycznie)
    
    ### Co robi aplikacja?
    
    - Wczytuje bazę inwestycji SPECTIS
    - Wykluczenia domyślne: Morskie elektrownie wiatrowe (z kolumny "Znaczące segmenty")
    - Wykluczenia opcjonalne: Twoja lista inwestycji
    - Tworzy 16 arkuszy wojewódzkich
    - Format: Sektor | Ogółem (mln zł) | W budowie (mln zł)
    
    **WAŻNE:** Wykluczanie inwestycji odbywa się po kolumnie **"Inwestycja"** (kolumna B).
    """)

st.markdown("---")

# KROK 1: Upload pliku
st.header("1️⃣ Wgraj plik Excel z bazą inwestycji")
uploaded_file = st.file_uploader(
    "Wybierz plik Excel (SPECTIS - baza inwestycji - *.xlsx)",
    type=['xlsx', 'xls'],
    help="Wgraj plik z arkuszem 'Inwestycje'"
)

# KROK 2: Wykluczenia (opcjonalne)
st.header("2️⃣ Inwestycje do wykluczenia (opcjonalne)")
wykluczenia_text = st.text_area(
    "Wpisz nazwy inwestycji do wykluczenia (po jednej w linii)",
    height=150,
    placeholder="Przykład:\nCPK - lotnisko\nElektrownia jądrowa Lubiatowo - elektrownia\nRail Baltica - odcinek polski\n\n# Linie zaczynające się od # są ignorowane (komentarze)",
    help="Wykluczanie odbywa się po dokładnej nazwie z kolumny 'Inwestycja' (kolumna B). Linie z # są traktowane jako komentarze."
)

# Przetworzenie listy wykluczeń (ignoruje linie z # i puste)
lista_wykluczanych = [
    nazwa.strip() 
    for nazwa in wykluczenia_text.strip().split('\n') 
    if nazwa.strip() and not nazwa.strip().startswith('#')
]

if lista_wykluczanych:
    st.info(f"✓ Wykluczysz {len(lista_wykluczanych)} inwestycji")
    with st.expander("Pokaż listę"):
        for i, nazwa in enumerate(lista_wykluczanych, 1):
            st.text(f"{i}. {nazwa}")

st.markdown("---")

# KROK 3: Przetwarzanie
st.header("3️⃣ Generowanie raportu")

if uploaded_file is not None:
    
    if st.button("🚀 Generuj raport", type="primary"):
        
        try:
            # Progress bar
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Wczytaj bazę
            status_text.text("📁 Wczytuję plik Excel...")
            progress_bar.progress(10)
            df = pd.read_excel(uploaded_file, sheet_name='Inwestycje')
            st.success(f"✓ Wczytano {len(df)} wierszy")
            
            # Filtrowanie
            status_text.text("🔍 Filtrowanie danych...")
            progress_bar.progress(20)
            
            # 1. Wykluczamy "Morskie elektrownie wiatrowe"
            df_filtered = df[~df['Znaczące segmenty'].fillna('').str.contains('Morskie elektrownie wiatrowe', case=False, na=False)].copy()
            
            # 2. Wykluczamy konkretne inwestycje
            if lista_wykluczanych:
                liczba_przed = len(df_filtered)
                df_filtered = df_filtered[~df_filtered['Inwestycja'].isin(lista_wykluczanych)]
                wykluczono = liczba_przed - len(df_filtered)
                st.info(f"✓ Wykluczono {wykluczono} wierszy z listy inwestycji")
            
            # 3. Zunifikuj nazwy województw (popraw literówki)
            df_filtered['Województwo'] = df_filtered['Województwo'].replace({
                'Kujawsko-Pomorskie': 'Kujawsko-pomorskie',
                'Warmińsko-Mazurskie': 'Warmińsko-mazurskie',
                'WIelkopolskie': 'Wielkopolskie'
            })
            
            # 4. Zunifikuj statusy
            df_filtered['Status inwestycji'] = df_filtered['Status inwestycji'].replace({
                'Wstępna Koncepcja': 'Wstępna koncepcja',
                'PLanowanie': 'Planowanie'
            })
            
            # 5. Zostaw tylko 4 statusy
            statusy_do_uwzglednienia = ['Budowa', 'Planowanie', 'Przetarg', 'Wstępna koncepcja']
            df_filtered = df_filtered[df_filtered['Status inwestycji'].isin(statusy_do_uwzglednienia)]
            
            progress_bar.progress(30)
            
            # 6. Konwersja wartości
            df_filtered['Wartość (mln zł)'] = pd.to_numeric(df_filtered['Wartość (mln zł)'], errors='coerce').fillna(0)
            
            # 7. Pobierz wszystkie sektory i województwa
            wszystkie_sektory = sorted(df_filtered['Sektor'].dropna().unique())
            wojewodztwa = sorted(df_filtered['Województwo'].unique(), key=polish_sort_key)
            
            st.success(f"✓ Znaleziono {len(wszystkie_sektory)} sektorów PKOB i {len(wojewodztwa)} województw")
            
            # Generowanie arkuszy
            status_text.text("📊 Generuję arkusze wojewódzkie...")
            progress_bar.progress(40)
            
            wb = Workbook()
            wb.remove(wb.active)
            
            for i, wojewodztwo in enumerate(wojewodztwa):
                
                # Update progress
                progress_value = 40 + int(50 * (i + 1) / len(wojewodztwa))
                progress_bar.progress(progress_value)
                status_text.text(f"📊 Generuję arkusz: {wojewodztwo} ({i+1}/{len(wojewodztwa)})")
                
                df_woj = df_filtered[df_filtered['Województwo'] == wojewodztwo]
                result_rows = []
                
                for sektor in wszystkie_sektory:
                    df_sektor = df_woj[df_woj['Sektor'] == sektor]
                    w_budowie = df_sektor[df_sektor['Status inwestycji'] == 'Budowa']['Wartość (mln zł)'].sum()
                    ogolem = df_sektor['Wartość (mln zł)'].sum()
                    
                    result_rows.append({
                        'Sektor': sektor,
                        'Ogółem (mln zł)': ogolem,
                        'W budowie (mln zł)': w_budowie
                    })
                
                result_df = pd.DataFrame(result_rows)
                result_df = result_df.sort_values('Ogółem (mln zł)', ascending=False)
                
                ws = wb.create_sheet(title=wojewodztwo[:31])
                
                for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        if r_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")
                
                ws.column_dimensions['A'].width = 80
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 20
            
            # Zapis do pamięci
            status_text.text("💾 Zapisuję plik...")
            progress_bar.progress(95)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            progress_bar.progress(100)
            status_text.text("✓ Gotowe!")
            
            # Informacje o pliku
            st.markdown("---")
            st.success("🎉 Raport został wygenerowany!")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Arkusze", len(wojewodztwa))
            with col2:
                st.metric("Sektory PKOB", len(wszystkie_sektory))
            with col3:
                st.metric("Wierszy danych", len(df_filtered))
            
            # Przycisk pobierania
            output_filename = f"wojewodztwa_inwestycje_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            
            st.download_button(
                label="📥 Pobierz raport",
                data=output,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
            
            # Podsumowanie
            with st.expander("📋 Szczegóły raportu"):
                st.markdown(f"""
                **Format wyjściowy:**
                - {len(wojewodztwa)} arkuszy (po 1 na województwo)
                - Kolumny: Sektor | Ogółem (mln zł) | W budowie (mln zł)
                - {len(wszystkie_sektory)} sektorów PKOB (z zerami gdzie brak danych)
                - Sortowanie od największych wartości
                
                **Filtry zastosowane:**
                - ❌ Wykluczono: Morskie elektrownie wiatrowe (z kolumny "Znaczące segmenty")
                {"- ❌ Wykluczono: " + str(len(lista_wykluczanych)) + " inwestycji z listy" if lista_wykluczanych else ""}
                - ✅ Statusy: Budowa, Planowanie, Przetarg, Wstępna koncepcja
                """)
            
        except Exception as e:
            st.error(f"❌ Wystąpił błąd: {str(e)}")
            st.exception(e)

else:
    st.info("👆 Zacznij od wgrania pliku Excel z bazą inwestycji")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray; font-size: 0.8em;'>
Generator Raportów Wojewódzkich SPECTIS | Styczeń 2025
</div>
""", unsafe_allow_html=True)
